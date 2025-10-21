from .Informe import Informe 
from models.interfaces.ExportarInformeInterfaz import ExportarInformeInterfaz
from fpdf import FPDF
from openpyxl import Workbook
from config import conectar_db
import datetime
import mysql.connector

class InformeGerente(Informe, ExportarInformeInterfaz):
    def __init__(self,
                  idInforme,
                  descripcion,
                  fecha,
                  rutGerente):
        super().__init__(idInforme, descripcion, fecha)

        self.rutGerente = rutGerente

    def __str__(self):
        return (f"Datos del informe de GERENTE:\n"
                f"ID: {self.idInforme}\n"
                f"Descripción: {self.descripcion}\n"
                f"Fecha: {self.fecha}\n"
                f"RUT Gerente: {self.rutGerente}")
    

    def formatearDatosParaExcel(self):
        return {
            "ID": self.idInforme,
            "Descripción": self.descripcion,
            "Fecha": self.fecha.strftime("%Y-%m-%d"),
            "RUT Gerente": self.rutGerente
        }
    
    """Métodos de la interfaz para exportar informes"""
    def generarPDF(self):
        class PDF(FPDF):
            def header(self):
                self.set_font("Arial", "B", 16)
                self.cell(0, 10, "Informe avances de tareas del Empleado", 0, 1, "C")

            def cuerpo(self):
                while True:
                    try:
                        id_informe = int(input("Ingrese la ID del informe que desea buscar: "))
                        break
                    except ValueError as Error:
                        print(f"Entrada inválida: {Error}")
                
                try:
                    conexion = conectar_db()
                    cursor = conexion.cursor(dictionary=True)

                    query = """
                        SELECT i.id_informe, i.descripcion, i.fecha, i.rut_usuario, u.rol
                        FROM informe i
                        JOIN usuario_detalle u ON i.rut_usuario = u.rut_usuario
                        WHERE i.id_informe = %s
                    """
                    cursor.execute(query, (id_informe,))
                    resultado = cursor.fetchone()

                    if not resultado:
                        print(f"No se encontró ningún informe con la ID {id_informe}.")
                        return None
                    
                    print("Informe encontrado con éxito.")
                    
                    #ID del informe
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "ID del informe:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.cell(0, 10, f"{resultado['id_informe']}", 0, 1, "C")
                    self.ln(5)

                    # RUT del usuario
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "RUT del usuario:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.cell(0, 10, f"{resultado['rut_usuario']}", 0, 1, "C")
                    self.ln(5)

                    # Rol
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "Rol del usuario:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.cell(0, 10, f"{resultado['rol'].capitalize()}", 0, 1, "C")
                    self.ln(5)

                    # Descripción
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "Descripción de avances:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.multi_cell(0, 10, f"{resultado['descripcion']}", 0, "C")
                    self.ln(5)

                    return id_informe

                except mysql.connector.Error as Error:
                    print(f"Error inesperado: {Error}")
                    return None
                finally:
                    if cursor:
                        cursor.close()
                    if conexion:
                        conexion.close()

            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 12)
                self.cell(0, 10, f"Generado el {datetime.datetime.now().strftime('%d/%m/%y')}", 0, 0, "C")

        # Crear el PDF
        archivo = PDF()
        archivo.add_page()
        id_pdf = archivo.cuerpo()

        if id_pdf:
            nombre_archivo = f"Informe_Empleado_{id_pdf}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            nombre_ruta = f"trabajoObjetos/docs/{nombre_archivo}"
            archivo.output(nombre_ruta)
            print(f"\nPDF generado exitosamente: {nombre_archivo}")
        else:
            print("\nNo se generó el PDF porque no se encontró el informe.")
    

    def generarExcel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Gerente"
        ws.append(["ID", "Descripción", "Fecha", "RUT Gerente"])
        ws.append([
            self.idInforme,
            self.descripcion,
            self.fecha.strftime("%Y-%m-%d"),
            self.rutGerente
        ])
        wb.save(f"informe_gerente{self.idInforme}.xlsx")

