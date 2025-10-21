from config import conectar_db
import mysql.connector
import re 
import datetime 
from fpdf import FPDF
from openpyxl import Workbook

def validar_rut(rut):
    rut = rut.strip().lower()
    if rut.count('-') != 1:
        raise ValueError("El RUT debe contener un solo guion ('-').")

    parte_num, dv = rut.split('-')

    if len(rut) < 9 or len(rut) > 10:
        raise ValueError("El RUT debe tener entre 9 y 10 caracteres en total.")

    if not parte_num.isdigit():
        raise ValueError("Los caracteres antes del guion deben ser solo números.")

    if len(parte_num) not in [7, 8]:
        raise ValueError("La parte numérica del RUT debe tener 7 u 8 dígitos.")

    if dv not in ['0','1','2','3','4','5','6','7','8','9','k']:
        raise ValueError("El dígito verificador debe ser un número o la letra 'k'.")

    return rut.upper()

def insertar_empleado_detalle(datos_detalle):

    try:
        conexion = conectar_db()
        cursor = conexion.cursor()
        query_detalle = """
            INSERT INTO usuario_detalle (
                rut_usuario, fecha_inicio_contrato,
                salario, rol, id_departamento
            ) VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query_detalle, datos_detalle)
        conexion.commit()
        print("Empleado creado con éxito.\n")
    except mysql.connector.Error as Error:
        print(f"Error inesperado: {Error}")
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

def insertar_empleado_completo(datos_basico, datos_detalle):
    try:
        conexion = conectar_db()
        cursor = conexion.cursor()
        query_basico = """
            INSERT INTO usuario_basico (
                rut_usuario, nombres, apellido_paterno, apellido_materno,
                fecha_nacimiento, numero_telefonico, direccion , contraseña, email
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        query_detalle = """
            INSERT INTO usuario_detalle (
                rut_usuario, fecha_inicio_contrato,
                salario, rol, id_departamento
            ) VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query_basico, datos_basico)
        cursor.execute(query_detalle, datos_detalle)
        conexion.commit()
        print("Empleado creado con éxito.\n")

    except mysql.connector.Error as Error:
        print(f"Error inesperado: {Error}")
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()




def validar_contraseña_segura(contraseña):
    if len(contraseña) < 8:
        raise ValueError("La contraseña debe tener al menos 8 caracteres.")
    
    if not re.search(r"[A-Z]", contraseña):
        raise ValueError("La contraseña debe contener al menos una letra mayúscula.")
    
    if not re.search(r"[a-z]", contraseña):
        raise ValueError("La contraseña debe contener al menos una letra minúscula.")
    
    if not re.search(r"\d", contraseña):
        raise ValueError("La contraseña debe contener al menos un número.")
    
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", contraseña):
        raise ValueError("La contraseña debe contener al menos un carácter especial.")
    
    return True

def buscar_empleado_general(rut):
    try:
        rut = validar_rut(rut)  # Si es válido, retorna el RUT limpio
    except ValueError as Error:
        print(Error)
        return  # Salir si el RUT no es válido

    try:
        conexion = conectar_db()
        cursor = conexion.cursor()

        query = """
        SELECT rut_usuario
        FROM usuario_basico
        WHERE rut_usuario = %s
        """
        cursor.execute(query, (rut,))
        resultado = cursor.fetchone()

        if resultado:
            print("El usuario se encuentra registrado en el sistema.\n")
            return True
        else:
            print("El usuario no está registrado.")
            return False
    except Exception as e:
        print(f"Error inesperado al buscar el empleado: {e}")
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

def buscar_proyecto_general(id_proyecto):
    try:
        conexion = conectar_db()
        cursor = conexion.cursor()

        query = """
                SELECT 
                    p.id_proyecto, 
                    p.nombre, 
                    p.descripcion, 
                    p.fecha_inicio, 
                    ud.rut_usuario, 
                    ud.id_departamento
                FROM 
                    proyecto AS p
                LEFT JOIN 
                    proyecto_has_usuario_detalle AS phu ON p.id_proyecto = phu.id_proyecto
                LEFT JOIN 
                    usuario_detalle AS ud ON phu.rut_usuario = ud.rut_usuario
                WHERE 
                    p.id_proyecto = %s;
                """
        cursor.execute(query, (id_proyecto,))
        resultado = cursor.fetchone()

        if resultado:
            print("El proyecto se encuentra registrado en el sistema.\n")
            return True
        else:
            print("No se encontró a ningún proyecto con esta ID.\n")
            return False
    except Exception as Error:
        print(f"Error inesperado al buscar el proyecto: {Error}")
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

def insertar_empleado_completo(datos_basico, datos_detalle):
    try:
        conexion = conectar_db()
        cursor = conexion.cursor()
        query_basico = """
            INSERT INTO usuario_basico (
                rut_usuario, nombres, apellido_paterno, apellido_materno,
                fecha_nacimiento, numero_telefonico, direccion, contraseña, email
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s,%s)
        """
        query_detalle = """
            INSERT INTO usuario_detalle (
                rut_usuario, fecha_inicio_contrato,
                salario, rol, id_departamento
            ) VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query_basico, datos_basico)
        cursor.execute(query_detalle, datos_detalle)
        conexion.commit()
        print("Empleado creado con éxito.\n")

    except mysql.connector.Error as Error:
        print(f"Error inesperado: {Error}")
    finally:
        if cursor:
            cursor.close()
        if conexion:
            conexion.close()

def generarPDF():
        class PDF(FPDF):
            def header(self):
                self.set_font("Arial", "B", 16)
                self.cell(0, 10, "Informe avances de tareas del Empleado", 0, 1, "C")

            def cuerpo(self):
                while True:
                    try:
                        id_informe = int(input("Ingrese la ID del informe que desea buscar: "))
                        break
                    except ValueError as Error:
                        print(f"Entrada inválida: {Error}")
                
                try:
                    conexion = conectar_db()
                    cursor = conexion.cursor(dictionary=True)

                    query = """
                        SELECT i.id_informe, i.descripcion, i.fecha, i.rut_usuario, u.rol
                        FROM informe i
                        JOIN usuario_detalle u ON i.rut_usuario = u.rut_usuario
                        WHERE i.id_informe = %s
                    """
                    cursor.execute(query, (id_informe,))
                    resultado = cursor.fetchone()

                    if not resultado:
                        print(f"No se encontró ningún informe con la ID {id_informe}.")
                        return None
                    
                    print("Informe encontrado con éxito.")
                    
                    #ID del informe
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "ID del informe:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.cell(0, 10, f"{resultado['id_informe']}", 0, 1, "C")
                    self.ln(5)

                    # RUT del usuario
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "RUT del usuario:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.cell(0, 10, f"{resultado['rut_usuario']}", 0, 1, "C")
                    self.ln(5)

                    # Rol
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "Rol del usuario:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.cell(0, 10, f"{resultado['rol'].capitalize()}", 0, 1, "C")
                    self.ln(5)

                    # Descripción
                    self.set_font("Arial", "B", 18)
                    self.cell(0, 10, "Descripción de avances:", 0, 1, "C")
                    self.set_font("Arial", "I", 12)
                    self.multi_cell(0, 10, f"{resultado['descripcion']}", 0, "C")
                    self.ln(5)

                    return id_informe

                except mysql.connector.Error as Error:
                    print(f"Error inesperado: {Error}")
                    return None
                finally:
                    if cursor:
                        cursor.close()
                    if conexion:
                        conexion.close()
            

            def footer(self):
                self.set_y(-15)
                self.set_font("Arial", "I", 12)
                self.cell(0, 10, f"Generado el {datetime.datetime.now().strftime('%d/%m/%y')}", 0, 0, "C")

        # Crear el PDF
        archivo = PDF()
        archivo.add_page()
        id_pdf = archivo.cuerpo()

        if id_pdf:
            nombre_archivo = f"Informe_Empleado_{id_pdf}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            nombre_ruta = f"trabajoObjetos/docs/{nombre_archivo}"
            archivo.output(nombre_ruta)
            print(f"\nPDF generado exitosamente: {nombre_archivo}")
        else:
            print("\nNo se generó el PDF porque no se encontró el informe.")

def generarExcel():
        while True:
            try:
                id_informe = int(input("Ingrese la ID del informe que desea buscar: "))
                break
            except ValueError as Error:
                print(f"Entrada inválida: {Error}")
                
        try:
            conexion = conectar_db()
            cursor = conexion.cursor(dictionary=True)

            query = """
                SELECT i.id_informe, i.descripcion, i.fecha, i.rut_usuario
                FROM informe i
                JOIN usuario_detalle u ON i.rut_usuario = u.rut_usuario
                WHERE i.id_informe = %s
            """
            cursor.execute(query, (id_informe,))
            resultado = cursor.fetchone()

            if not resultado:
                print(f"No se encontró ningún informe con la ID {id_informe}.")
                return None
                    
            print("Informe encontrado con éxito.")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe Gerente"

            # Encabezados
            ws.append(["ID", "Descripción", "Fecha", "RUT Gerente"])

            # Fila con datos
            ws.append([
                resultado['id_informe'],
                resultado['descripcion'],
                resultado['fecha'].strftime("%Y-%m-%d") if resultado['fecha'] else "Sin fecha",
                resultado['rut_usuario'],
            ])

            nombre_archivo = f"Informe_Gerente_{id_informe}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta_archivo = f"trabajoObjetos/docs/{nombre_archivo}"
            wb.save(ruta_archivo)
            print(f"Archivo Excel generado correctamente: {nombre_archivo}")

        except mysql.connector.Error as Error:
            print(f"Error inesperado: {Error}")

        finally:
            if cursor:
                cursor.close()
            if conexion:
                conexion.close()